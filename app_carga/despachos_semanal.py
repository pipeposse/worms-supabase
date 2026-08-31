# -*- coding: utf-8 -*-
"""worms_supabase / app_carga / despachos_semanal.py

Despachos → vista "⬇️ Semanal": descarga rápida de los despachos de cada semana con sus
materias primas en TN, como Excel (una hoja por semana + resumen + detalle de líneas) y
PNG (una lámina por semana con gráfico apilado + tabla; más una lámina resumen).

Fuente: fact_despacho × fact_despacho_linea (lo FORMULADO). TN = litros × densidad de la
línea (la misma regla que usa el armador). Se agrupa por producto de la línea (materia
prima / diluyente) y por semana ISO del despacho.

AJUSTE A LOS KG REALES
----------------------
Lo formulado casi nunca es lo que sale: contra los tickets de portería los despachos dan
2.000 a 5.000 kg de diferencia. Pero la balanza pesa por CONTENEDOR y no dice de qué
tanque cargó cada camión, así que no se puede repartir sola por materia prima — eso lo
sabe quien despachó.

Por eso la tabla de la semana se puede EDITAR: se escriben las TN reales de cada materia
prima y se guardan en `produccion.fact_despacho_real`, que **no pisa la formulación**
(misma lógica PLANIFICADO → EJECUTADO que usa el stock). De ahí en más todo — tabla,
resumen, Excel y PNG — muestra los kg EFECTIVOS: el real cuando alguien lo ajustó y el
formulado cuando no, vía `produccion.v_despacho_mp`. El control contra la balanza sale de
`produccion.v_despacho_real_vs_ticket`.
"""

import io
import zipfile
from datetime import date, timedelta

import pandas as pd
import streamlit as st

# Grilla propia para el ajuste (ver ajuste_mp/): un st.data_editor vuelve al servidor en
# cada celda y redibuja la sección entera — en planta eso se siente como que "se cuelga y
# se reinicia". Si el componente no carga, queda el data_editor como respaldo.
try:
    from ajuste_mp import grilla as _grilla_mp, disponible as _grilla_ok
except Exception:
    _grilla_mp = None
    _grilla_ok = lambda: False

ESTADOS_DEF = ["CONFIRMADO", "DESPACHADO"]
# colores fijos por materia prima para que todas las láminas se lean igual
COLORES = {"AFE-S": "#2563eb", "AG-E": "#f59e0b", "AFE-AL": "#10b981", "AFE-M": "#8b5cf6",
           "AFE-G": "#14b8a6", "AFE-P": "#0ea5e9", "AG-C": "#ef4444", "AG-A": "#f97316",
           "AG-B": "#fb7185", "ARE-A": "#a16207", "ARE-B": "#78350f", "ARE-A-ANIMAL": "#92400e"}
_PALETA_EXTRA = ["#64748b", "#84cc16", "#e11d48", "#06b6d4", "#a855f7", "#eab308"]


def _color(mp, i):
    return COLORES.get(str(mp), _PALETA_EXTRA[i % len(_PALETA_EXTRA)])


def _sem_label(anio, sem):
    return "%d-S%02d" % (int(anio), int(sem))


def _sem_rango(anio, sem):
    try:
        d1 = date.fromisocalendar(int(anio), int(sem), 1)
        d2 = d1 + timedelta(days=6)
        return "%s–%s" % (d1.strftime("%d/%m"), d2.strftime("%d/%m"))
    except Exception:
        return ""


# ------------------------------------------------------------------ datos

def _lineas(cat, estados, dias):
    return cat(
        "SELECT d.id_despacho, d.titulo, COALESCE(NULLIF(d.cliente,''),'EGNITRADE') AS cliente, d.destino, "
        "       d.producto_codigo AS producto_venta, d.tipo_carga, d.estado, d.fecha_despacho, "
        "       COALESCE(d.anio, EXTRACT(isoyear FROM d.fecha_despacho))::int AS anio, "
        "       COALESCE(d.semana_iso, EXTRACT(week FROM d.fecha_despacho))::int AS semana_iso, "
        "       COALESCE(d.n_contenedores,0) AS n_contenedores, "
        "       l.orden, l.producto_codigo AS mp, t.nombre AS tanque, l.litros, "
        "       COALESCE(l.densidad, p.densidad_g_ml, 0.9) AS densidad, "
        "       l.litros * COALESCE(l.densidad, p.densidad_g_ml, 0.9) AS kg "
        "FROM produccion.fact_despacho d "
        "JOIN produccion.fact_despacho_linea l ON l.id_despacho = d.id_despacho "
        "LEFT JOIN produccion.dim_tanque t ON t.id_tanque = l.id_tanque "
        "LEFT JOIN produccion.dim_producto p ON p.codigo_producto = l.producto_codigo "
        "WHERE d.estado = ANY(%s) AND d.fecha_despacho >= current_date - %s "
        "ORDER BY d.fecha_despacho, d.id_despacho, l.orden", (list(estados), int(dias)))


def _ajustes(cat):
    """Kg REALES cargados a mano, por despacho y materia prima (los que haya)."""
    df = cat("SELECT id_despacho, producto_codigo AS mp, kg_real, ajustado, "
             "       kg_formulado, nota "
             "FROM produccion.v_despacho_mp WHERE ajustado")
    if df is None or df.empty:
        return pd.DataFrame(columns=["id_despacho", "mp", "kg_real", "ajustado",
                                     "kg_formulado", "nota"])
    df = df.copy()
    for c in ("kg_real", "kg_formulado"):
        df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0.0)
    df["id_despacho"] = pd.to_numeric(df["id_despacho"], errors="coerce").fillna(0).astype(int)
    return df


def _tickets(cat, dias):
    """Kg pesados en portería por despacho: el control de que el ajuste cierre."""
    df = cat("SELECT id_despacho, kg_tickets, n_tickets, n_contenedores "
             "FROM produccion.v_despacho_real_vs_ticket "
             "WHERE fecha_despacho >= current_date - %s", (int(dias),))
    if df is None or df.empty:
        return pd.DataFrame(columns=["id_despacho", "kg_tickets", "n_tickets", "n_contenedores"])
    df = df.copy()
    for c in ("kg_tickets", "n_tickets", "n_contenedores"):
        df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0.0)
    df["id_despacho"] = pd.to_numeric(df["id_despacho"], errors="coerce").fillna(0).astype(int)
    return df


def _aplicar_ajustes(df, aj):
    """Reemplaza los kg formulados por los REALES donde alguien los haya ajustado.

    El ajuste se carga por materia prima y la tabla de líneas es por TANQUE, así que
    los kg reales se prorratean entre las líneas de esa materia prima (si la MP tenía
    0 L formulados —entró algo que no estaba en la mezcla— se agrega una fila propia).
    Los litros se recalculan con la densidad de cada línea para que el detalle cierre.
    """
    df = df.copy()
    df["ajustado"] = False
    if aj is None or aj.empty:
        return df
    idx = {(int(r["id_despacho"]), str(r["mp"])): float(r["kg_real"])
           for _, r in aj.iterrows()}
    if not idx:
        return df
    nuevas = []
    for (idd, mp), kg_real in idx.items():
        m = (df["id_despacho"] == idd) & (df["mp"] == mp)
        kg_f = float(df.loc[m, "kg"].sum())
        if m.any() and kg_f > 0.5:
            df.loc[m, "kg"] = df.loc[m, "kg"] * (kg_real / kg_f)   # prorrateo por línea
            df.loc[m, "litros"] = df.loc[m, "kg"] / df.loc[m, "densidad"].replace(0, pd.NA)
            df.loc[m, "ajustado"] = True
        elif m.any():
            # estaba formulada en 0: se le da todo a la primera línea
            _i = df.index[m][0]
            df.loc[_i, "kg"] = kg_real
            df.loc[m, "ajustado"] = True
        else:
            # materia prima que NO estaba formulada y sí entró: fila sintética
            _cab = df[df["id_despacho"] == idd]
            if _cab.empty or kg_real <= 0:
                continue
            _r = _cab.iloc[0].to_dict()
            _r.update({"mp": mp, "tanque": None, "orden": 99, "kg": kg_real,
                       "densidad": 0.9, "litros": kg_real / 0.9, "ajustado": True})
            nuevas.append(_r)
    if nuevas:
        df = pd.concat([df, pd.DataFrame(nuevas)], ignore_index=True)
    df["kg"] = pd.to_numeric(df["kg"], errors="coerce").fillna(0.0)
    df["litros"] = pd.to_numeric(df["litros"], errors="coerce").fillna(0.0)
    return df


def _guardar_titulos(conectar, USR, renombres):
    """Renombra despachos desde la grilla. renombres = [(id_despacho, titulo)].

    El título es de fact_despacho: renombrar acá lo renombra en TODA la app (armador,
    listado, control, brief). Queda auditado con usuario y fecha."""
    if not renombres:
        return 0
    n = 0
    with conectar(int(USR["id_usuario"])) as (conn, audit):
        with conn.cursor() as cur:
            for idd, tit in renombres:
                tit = str(tit or "").strip()[:80]
                if not tit:
                    continue
                cur.execute("UPDATE produccion.fact_despacho "
                            "SET titulo=%s, actualizado_en=now() WHERE id_despacho=%s",
                            (tit, int(idd)))
                audit.log("U", "fact_despacho", int(idd),
                          {"titulo": tit, "desde": "semanal"})
                n += 1
    return n


def _guardar_ajustes(conectar, USR, cambios):
    """Upsert de los kg reales. cambios = [(id_despacho, mp, kg_real|None)].

    kg_real None borra el ajuste (esa materia prima vuelve a valer lo formulado)."""
    if not cambios:
        return 0
    n = 0
    with conectar(int(USR["id_usuario"])) as (conn, audit):
        with conn.cursor() as cur:
            for idd, mp, kg in cambios:
                if kg is None:
                    cur.execute("DELETE FROM produccion.fact_despacho_real "
                                "WHERE id_despacho=%s AND producto_codigo=%s",
                                (int(idd), str(mp)))
                else:
                    cur.execute(
                        "INSERT INTO produccion.fact_despacho_real "
                        "(id_despacho, producto_codigo, kg_real, id_usuario, actualizado_en) "
                        "VALUES (%s,%s,%s,%s,now()) "
                        "ON CONFLICT (id_despacho, producto_codigo) DO UPDATE SET "
                        " kg_real=EXCLUDED.kg_real, id_usuario=EXCLUDED.id_usuario, "
                        " actualizado_en=now()",
                        (int(idd), str(mp), round(float(kg), 2), int(USR["id_usuario"])))
                n += 1
        audit.log("U", "fact_despacho_real", int(cambios[0][0]),
                  {"ajustes": [[int(a), str(b), (None if c is None else round(float(c), 2))]
                               for a, b, c in cambios]})
    return n


def _preparar(df):
    df = df.copy()
    for c in ("litros", "densidad", "kg", "n_contenedores"):
        df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0.0)
    df["tn"] = df["kg"] / 1000.0
    df["mp"] = df["mp"].fillna("?").astype(str)
    df["Semana"] = [_sem_label(a, s) for a, s in zip(df["anio"], df["semana_iso"])]
    df["Rango"] = [_sem_rango(a, s) for a, s in zip(df["anio"], df["semana_iso"])]
    df["fecha_despacho"] = pd.to_datetime(df["fecha_despacho"], errors="coerce")
    return df


def _orden_mp(df):
    """AFE-S primero (es el grueso), después el resto por TN totales."""
    tot = df.groupby("mp")["tn"].sum().sort_values(ascending=False)
    mps = [m for m in tot.index if m == "AFE-S"] + [m for m in tot.index if m != "AFE-S"]
    return mps


def _tabla_semana(df_s, mps):
    """Despachos de una semana × materias primas (TN) + fila TOTAL."""
    cab = (df_s.groupby("id_despacho", sort=False)
               .agg(Despacho=("titulo", "first"), Fecha=("fecha_despacho", "first"),
                    Producto=("producto_venta", "first"), Cliente=("cliente", "first"),
                    Destino=("destino", "first"), Cont=("n_contenedores", "first"))
               .reset_index())
    piv = df_s.pivot_table(index="id_despacho", columns="mp", values="tn", aggfunc="sum", fill_value=0.0)
    piv = piv.reindex(columns=[m for m in mps if m in piv.columns], fill_value=0.0)
    piv["TN total"] = piv.sum(axis=1)
    t = cab.merge(piv.reset_index(), on="id_despacho", how="left").sort_values(["Fecha", "id_despacho"])
    t["Fecha"] = t["Fecha"].dt.strftime("%d/%m/%Y")
    # marca de "kg reales": se ve en la tabla HTML, y las descargas la ignoran
    if "ajustado" in df_s.columns:
        _aj = df_s.groupby("id_despacho")["ajustado"].any()
        t["_aj"] = t["id_despacho"].map(_aj).fillna(False)
    else:
        t["_aj"] = False
    t = t.drop(columns=["id_despacho"])
    mp_cols = [c for c in t.columns if c in mps] + ["TN total"]
    tot = {c: "" for c in t.columns}
    tot["_aj"] = False
    tot["Despacho"] = "TOTAL SEMANA"
    tot["Cont"] = int(t["Cont"].sum())
    for c in mp_cols:
        tot[c] = float(t[c].sum())
    t = pd.concat([t, pd.DataFrame([tot])], ignore_index=True)
    for c in mp_cols:
        t[c] = pd.to_numeric(t[c], errors="coerce").round(2)
    t["Cont"] = pd.to_numeric(t["Cont"], errors="coerce").fillna(0).astype(int)
    return t, mp_cols


def _tabla_resumen(df, mps):
    """Semanas × materias primas (TN) + despachos + contenedores + TN total."""
    piv = df.pivot_table(index=["Semana", "Rango"], columns="mp", values="tn", aggfunc="sum", fill_value=0.0)
    piv = piv.reindex(columns=[m for m in mps if m in piv.columns], fill_value=0.0)
    piv["TN total"] = piv.sum(axis=1)
    n = df.groupby(["Semana", "Rango"]).agg(Despachos=("id_despacho", "nunique"))
    cont = (df.drop_duplicates("id_despacho").groupby(["Semana", "Rango"])["n_contenedores"].sum()
              .rename("Contenedores"))
    r = n.join(cont).join(piv).reset_index().sort_values("Semana")
    tot = {"Semana": "TOTAL", "Rango": "", "Despachos": int(r["Despachos"].sum()),
           "Contenedores": int(r["Contenedores"].sum())}
    for c in [m for m in mps if m in r.columns] + ["TN total"]:
        tot[c] = float(r[c].sum())
    r = pd.concat([r, pd.DataFrame([tot])], ignore_index=True)
    for c in [m for m in mps if m in r.columns] + ["TN total"]:
        r[c] = pd.to_numeric(r[c], errors="coerce").round(2)
    for c in ("Despachos", "Contenedores"):
        r[c] = pd.to_numeric(r[c], errors="coerce").fillna(0).astype(int)
    return r


# ------------------------------------------------------------------ Excel

def _excel(df, mps):
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    buf = io.BytesIO()
    semanas = sorted(df["Semana"].unique().tolist(), reverse=True)
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        res = _tabla_resumen(df, mps)
        res.to_excel(w, sheet_name="Resumen semanal", index=False)
        hojas = [("Resumen semanal", res, ["TN total"] + [m for m in mps if m in res.columns])]
        for s in semanas:
            df_s = df[df["Semana"] == s]
            t, mp_cols = _tabla_semana(df_s, mps)
            t = t.drop(columns=[c for c in ("_aj",) if c in t.columns])
            nombre = ("%s %s" % (s, df_s["Rango"].iloc[0])).replace("/", "-")[:31]
            t.to_excel(w, sheet_name=nombre, index=False)
            hojas.append((nombre, t, mp_cols))
        _cols_det = ["Semana", "Rango", "fecha_despacho", "titulo", "producto_venta", "cliente",
                     "destino", "estado", "n_contenedores", "orden", "mp", "tanque", "litros",
                     "densidad", "tn"]
        if "ajustado" in df.columns:
            _cols_det.append("ajustado")
        det = df[_cols_det].copy()
        if "ajustado" in det.columns:
            det["ajustado"] = det["ajustado"].map({True: "kg reales", False: "formulado"}).fillna("formulado")
        det["fecha_despacho"] = det["fecha_despacho"].dt.strftime("%d/%m/%Y")
        det = det.rename(columns={"fecha_despacho": "Fecha", "titulo": "Despacho", "producto_venta": "Producto venta",
                                  "cliente": "Cliente", "destino": "Destino", "estado": "Estado",
                                  "n_contenedores": "Cont", "orden": "Línea", "mp": "Materia prima",
                                  "tanque": "Tanque", "litros": "Litros", "densidad": "Densidad",
                                  "tn": "TN", "ajustado": "Origen"})
        det["TN"] = det["TN"].round(3)
        det.to_excel(w, sheet_name="Detalle líneas", index=False)
        hojas.append(("Detalle líneas", det, ["TN"]))

        head_fill = PatternFill("solid", fgColor="1e3a8a")
        tot_fill = PatternFill("solid", fgColor="e0e7ff")
        for nombre, t, num_cols in hojas:
            ws = w.sheets[nombre]
            for c in range(1, len(t.columns) + 1):
                cell = ws.cell(row=1, column=c)
                cell.font = Font(bold=True, color="ffffff")
                cell.fill = head_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                col = t.columns[c - 1]
                if col in num_cols:
                    for r in range(2, len(t) + 2):
                        ws.cell(row=r, column=c).number_format = "#,##0.00"
                elif col in ("Litros",):
                    for r in range(2, len(t) + 2):
                        ws.cell(row=r, column=c).number_format = "#,##0"
                ancho = max([len(str(col))] + [len(str(v)) for v in t[col].head(200).tolist()]) + 2
                ws.column_dimensions[get_column_letter(c)].width = min(max(ancho, 8), 38)
            if nombre != "Detalle líneas":
                for c in range(1, len(t.columns) + 1):
                    cell = ws.cell(row=len(t) + 1, column=c)
                    cell.font = Font(bold=True)
                    cell.fill = tot_fill
            ws.freeze_panes = "B2"
    return buf.getvalue()


# ------------------------------------------------------------------ PNG

def _plt():
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    plt.rcParams.update({"font.family": "DejaVu Sans", "font.size": 9, "axes.spines.top": False,
                         "axes.spines.right": False})
    return plt


def _fmt(v):
    return "{:,.2f}".format(float(v)).replace(",", "@").replace(".", ",").replace("@", ".")


def _png_semana(df_s, mps, semana, rango):
    plt = _plt()
    t, mp_cols = _tabla_semana(df_s, mps)
    cuerpo = t.iloc[:-1]
    total = t.iloc[-1]
    mp_use = [m for m in mp_cols if m != "TN total" and float(t[m].iloc[-1]) > 0]
    n = len(cuerpo)
    alto_bar = max(2.2, 0.55 * n + 1.2)
    alto_tab = max(1.2, 0.34 * (n + 2))
    fig = plt.figure(figsize=(12, alto_bar + alto_tab + 1.1), dpi=150)
    gs = fig.add_gridspec(2, 1, height_ratios=[alto_bar, alto_tab], hspace=0.35)
    ax = fig.add_subplot(gs[0])
    fig.suptitle("Despachos semana %s  ·  %s  ·  %d despachos  ·  %d contenedores  ·  %s TN" % (
        semana, rango, n, int(total["Cont"]), _fmt(total["TN total"])),
        fontsize=12.5, fontweight="bold", x=0.02, ha="left", y=0.995)

    y = list(range(n))[::-1]
    izq = [0.0] * n
    for i, m in enumerate(mp_use):
        vals = cuerpo[m].astype(float).tolist()
        ax.barh(y, vals, left=izq, color=_color(m, i), label=m, height=0.62, edgecolor="white", linewidth=0.5)
        for yi, v, l in zip(y, vals, izq):
            if v >= max(3.0, 0.06 * float(total["TN total"] or 1)):
                ax.text(l + v / 2, yi, _fmt(v), ha="center", va="center", fontsize=7.5, color="white",
                        fontweight="bold")
        izq = [a + b for a, b in zip(izq, vals)]
    for yi, tot_i in zip(y, cuerpo["TN total"].astype(float).tolist()):
        ax.text(tot_i + 0.5, yi, _fmt(tot_i) + " TN", va="center", fontsize=8, fontweight="bold", color="#111827")
    ax.set_yticks(y)
    ax.set_yticklabels(["%s  (%s · %d cont.)" % (r["Despacho"], r["Fecha"][:5], int(r["Cont"]))
                        for _, r in cuerpo.iterrows()], fontsize=8.5)
    ax.set_xlabel("TN formuladas por materia prima")
    ax.set_xlim(0, max(1.0, float(cuerpo["TN total"].max()) * 1.18))
    ax.grid(axis="x", alpha=0.25)
    ax.legend(loc="upper center", bbox_to_anchor=(0.5, -0.16 - 0.9 / max(n, 1) * 0.3), ncol=min(len(mp_use), 6),
              frameon=False, fontsize=8)

    axt = fig.add_subplot(gs[1])
    axt.axis("off")
    cols = ["Despacho", "Fecha", "Producto", "Cliente", "Cont"] + mp_use + ["TN total"]
    celdas = []
    for _, r in t.iterrows():
        fila = []
        for c in cols:
            v = r[c]
            if c in mp_use or c == "TN total":
                fila.append(_fmt(v) if float(v or 0) else "—")
            elif c == "Cont":
                fila.append(str(int(v)) if str(v) != "" else "")
            else:
                fila.append(str(v)[:22])
        celdas.append(fila)
    tb = axt.table(cellText=celdas, colLabels=cols, loc="upper center", cellLoc="center")
    tb.auto_set_font_size(False)
    tb.set_fontsize(7.6)
    tb.scale(1, 1.25)
    for (ri, ci), cell in tb.get_celld().items():
        cell.set_edgecolor("#d1d5db")
        if ri == 0:
            cell.set_facecolor("#1e3a8a")
            cell.set_text_props(color="white", fontweight="bold")
        elif ri == len(celdas):
            cell.set_facecolor("#e0e7ff")
            cell.set_text_props(fontweight="bold")
        elif ri % 2 == 0:
            cell.set_facecolor("#f8fafc")
    fig.text(0.02, 0.005, "Fuente: formulación de despachos (litros × densidad por línea). Estados: %s." % (
        ", ".join(sorted(df_s["estado"].unique().tolist()))), fontsize=7, color="#6b7280")
    buf = io.BytesIO()
    fig.savefig(buf, format="png", bbox_inches="tight", facecolor="white")
    plt.close(fig)
    return buf.getvalue()


def _png_resumen(df, mps):
    plt = _plt()
    r = _tabla_resumen(df, mps).iloc[:-1]
    mp_use = [m for m in mps if m in r.columns and float(r[m].sum()) > 0]
    n = len(r)
    fig, ax = plt.subplots(figsize=(max(8, 0.9 * n + 4), 5.2), dpi=150)
    x = list(range(n))
    base = [0.0] * n
    for i, m in enumerate(mp_use):
        vals = r[m].astype(float).tolist()
        ax.bar(x, vals, bottom=base, color=_color(m, i), label=m, width=0.66, edgecolor="white", linewidth=0.5)
        for xi, v, b in zip(x, vals, base):
            if v >= max(5.0, 0.07 * float(r["TN total"].max() or 1)):
                ax.text(xi, b + v / 2, _fmt(v), ha="center", va="center", fontsize=7.5, color="white",
                        fontweight="bold")
        base = [a + b for a, b in zip(base, vals)]
    for xi, tot, nd in zip(x, r["TN total"].astype(float).tolist(), r["Despachos"].tolist()):
        ax.text(xi, tot + float(r["TN total"].max()) * 0.015, "%s TN\n%d desp." % (_fmt(tot), int(nd)),
                ha="center", va="bottom", fontsize=8, fontweight="bold", color="#111827")
    ax.set_xticks(x)
    ax.set_xticklabels(["%s\n%s" % (s, g) for s, g in zip(r["Semana"], r["Rango"])], fontsize=8)
    ax.set_ylabel("TN formuladas")
    ax.set_ylim(0, float(r["TN total"].max() or 1) * 1.22)
    ax.grid(axis="y", alpha=0.25)
    ax.legend(loc="upper left", ncol=min(len(mp_use), 6), frameon=False, fontsize=8)
    ax.set_title("Despachos por semana y materia prima (TN)  ·  total %s TN en %d despachos" % (
        _fmt(r["TN total"].sum()), int(r["Despachos"].sum())), fontsize=12, fontweight="bold", loc="left")
    buf = io.BytesIO()
    fig.savefig(buf, format="png", bbox_inches="tight", facecolor="white")
    plt.close(fig)
    return buf.getvalue()


@st.cache_data(ttl=300, show_spinner=False)
def _paquete(df, mps):
    """Excel + todos los PNG, comprimidos. Cacheado: se arma una vez por combinación de datos."""
    xlsx = _excel(df, mps)
    zbuf = io.BytesIO()
    with zipfile.ZipFile(zbuf, "w", zipfile.ZIP_DEFLATED) as z:
        z.writestr("despachos_semanal.xlsx", xlsx)
        z.writestr("resumen_semanas.png", _png_resumen(df, mps))
        for s in sorted(df["Semana"].unique().tolist()):
            df_s = df[df["Semana"] == s]
            z.writestr("despachos_%s.png" % s, _png_semana(df_s, mps, s, df_s["Rango"].iloc[0]))
    return xlsx, zbuf.getvalue()


# ------------------------------------------------------------------ UI

def _filas_grilla(df_s, mps, tk, aj):
    """Las filas que consume el componente: efectivo, formulado y balanza por despacho."""
    cab = (df_s.groupby("id_despacho", sort=False)
               .agg(desp=("titulo", "first"), fecha=("fecha_despacho", "first"),
                    cont=("n_contenedores", "first"))
               .reset_index())
    ef = (df_s.groupby(["id_despacho", "mp"])["tn"].sum().unstack(fill_value=0.0)
          if not df_s.empty else pd.DataFrame())
    _ajd = {}
    if aj is not None and not aj.empty:
        for _, r in aj.iterrows():
            _ajd.setdefault(int(r["id_despacho"]), {})[str(r["mp"])] = float(r["kg_formulado"]) / 1000.0
    _bal = {}
    if tk is not None and not tk.empty:
        _bal = {int(r["id_despacho"]): float(r["kg_tickets"]) / 1000.0 for _, r in tk.iterrows()}
    _ajado = (df_s.groupby("id_despacho")["ajustado"].any().to_dict()
              if "ajustado" in df_s.columns else {})
    filas = []
    for _, c in cab.iterrows():
        idd = int(c["id_despacho"])
        v, f = {}, {}
        for m in mps:
            _e = float(ef.loc[idd, m]) if (idd in ef.index and m in ef.columns) else 0.0
            v[m] = round(_e, 2)
            # el formulado real sale de v_despacho_mp cuando la MP está ajustada;
            # si no está ajustada, el efectivo ES el formulado
            f[m] = round(_ajd.get(idd, {}).get(m, _e), 2)
        try:
            _fe = pd.to_datetime(c["fecha"]).strftime("%d/%m")
        except Exception:
            _fe = ""
        filas.append({"id": idd, "desp": str(c["desp"]), "fecha": _fe,
                      "cont": int(c["cont"] or 0), "bal": round(_bal.get(idd, 0.0), 2),
                      "aj": bool(_ajado.get(idd, False)), "v": v, "f": f})
    return filas


def _grilla_reales(df_s, mps, tk, aj, cat, conectar, USR, semana):
    """Ajuste de kg reales en el componente propio: instantáneo y sin reruns.

    Devuelve True si lo dibujó, False si hay que caer al editor clásico.
    """
    if not _grilla_ok():
        return False
    ss = st.session_state
    st.caption("Lo formulado **no se pisa**: el ajuste se guarda aparte con tu usuario y fecha. "
               "**Balanza** es lo que pesó portería y **⚖** reparte la diferencia para que el "
               "total dé exactamente eso. El **nombre del despacho** también se edita acá y "
               "se renombra en toda la app.")
    _rev = int(ss.get("dsw_rev") or 0)
    # la confirmación se muestra DESPUÉS del rerun (un st.success antes de st.rerun
    # aparece medio parpadeo y desaparece: nadie llega a leerlo)
    _msg_ok = ss.pop("dsw_msg", None)
    if _msg_ok:
        st.success(_msg_ok)
    val = _grilla_mp(rows=_filas_grilla(df_s, mps, tk, aj), mps=mps, rev=_rev,
                     titulo="Semana %s" % semana, key="dsw_grilla")
    if isinstance(val, dict) and val.get("action") == "save" \
            and int(val.get("rev", -1)) == _rev \
            and int(val.get("seq", 0)) != int(ss.get("dsw_seq") or 0):
        ss["dsw_seq"] = int(val.get("seq", 0))
        cambios = []
        for c in (val.get("cambios") or []):
            try:
                _tn = c.get("tn")
                cambios.append((int(c["id"]), str(c["mp"]),
                                None if _tn is None else float(_tn) * 1000.0))
            except Exception:
                continue
        renombres = []
        for c in (val.get("titulos") or []):
            try:
                renombres.append((int(c["id"]), str(c.get("titulo") or "")))
            except Exception:
                continue
        if cambios or renombres:
            if conectar is None:
                st.error("Esta vista se abrió sin permisos de escritura.")
            else:
                try:
                    n = _guardar_ajustes(conectar, USR, cambios) if cambios else 0
                    nt = _guardar_titulos(conectar, USR, renombres) if renombres else 0
                    cat.clear()
                    ss["dsw_rev"] = _rev + 1      # rev nueva: la grilla se resincroniza
                    _msg = []
                    if n:
                        _msg.append("%d ajuste(s) de kg" % n)
                    if nt:
                        _msg.append("%d despacho(s) renombrado(s)" % nt)
                    ss["dsw_msg"] = ("Guardado: %s. La tabla, el Excel y los PNG ya lo "
                                     "muestran." % " y ".join(_msg))
                    st.rerun()
                except Exception as e:
                    st.error("No se pudo guardar: %s" % e)
    if st.checkbox("🧱 Usar la tabla clásica de Streamlit en vez de esta grilla",
                   key="dsw_clasico",
                   help="Respaldo por si la grilla no funciona en alguna máquina de planta. "
                        "La tabla clásica vuelve al servidor en cada celda y es más lenta."):
        return "clasico"
    return True


def _editor_reales(df_s, mps, tk, cat, conectar, USR, semana):
    """Grilla editable: TN REALES de cada materia prima, despacho por despacho.

    Se edita en TN (lo que se ve) y se guarda en kg. Al lado de la mezcla van los kg que
    pesó la balanza para ese despacho: si el ajuste no cierra contra portería, se ve al
    instante y no hace falta ir a buscarlo a otra pantalla.
    """
    st.markdown("##### ✏️ Ajustar a los kg realmente cargados")
    st.caption("Escribí las **TN reales** de cada materia prima. Lo formulado **no se pisa**: "
               "el ajuste se guarda aparte con tu usuario y fecha, y de ahí en más la tabla, "
               "el Excel y los PNG muestran este número. Dejá una celda **vacía** para que esa "
               "materia prima vuelva a valer lo formulado. La columna **Balanza** es lo que "
               "pesó portería: es contra eso que tiene que cerrar el total.")

    cab = (df_s.groupby("id_despacho", sort=False)
               .agg(Despacho=("titulo", "first"), Fecha=("fecha_despacho", "first"),
                    Cliente=("cliente", "first"), Cont=("n_contenedores", "first"))
               .reset_index())
    piv_ef = df_s.pivot_table(index="id_despacho", columns="mp", values="tn",
                              aggfunc="sum", fill_value=0.0)
    piv_ef = piv_ef.reindex(columns=[m for m in mps if m in piv_ef.columns], fill_value=0.0)
    _cols_mp = list(piv_ef.columns)
    _aj = (df_s.groupby("id_despacho")["ajustado"].any()
           if "ajustado" in df_s.columns else None)

    base = cab.merge(piv_ef.reset_index(), on="id_despacho", how="left").fillna(0.0)
    base = base.merge(tk[["id_despacho", "kg_tickets", "n_tickets"]], on="id_despacho",
                      how="left")
    base["Balanza (TN)"] = pd.to_numeric(base["kg_tickets"], errors="coerce").fillna(0.0) / 1000.0
    base["Fecha"] = pd.to_datetime(base["Fecha"], errors="coerce").dt.strftime("%d/%m")
    base["Aj."] = [("✏️ real" if (_aj is not None and bool(_aj.get(i, False))) else "· form.")
                   for i in base["id_despacho"]]
    base["Mezcla (TN)"] = base[_cols_mp].sum(axis=1).round(2)
    base["Dif. vs balanza"] = (base["Mezcla (TN)"] - base["Balanza (TN)"]).round(2)
    for c in _cols_mp + ["Mezcla (TN)", "Balanza (TN)", "Dif. vs balanza"]:
        base[c] = pd.to_numeric(base[c], errors="coerce").round(2)

    _ids = base["id_despacho"].tolist()
    _vista = base[["Aj.", "Despacho", "Fecha", "Cont"] + _cols_mp +
                  ["Mezcla (TN)", "Balanza (TN)", "Dif. vs balanza"]].copy()

    _cfg = {"Aj.": st.column_config.TextColumn(
                "Aj.", width="small", disabled=True,
                help="✏️ real = este despacho ya tiene kg reales cargados · "
                     "· form. = todavía muestra lo formulado."),
            "Despacho": st.column_config.TextColumn(disabled=True),
            "Fecha": st.column_config.TextColumn(disabled=True, width="small"),
            "Cont": st.column_config.NumberColumn("Cont.", disabled=True, width="small"),
            "Mezcla (TN)": st.column_config.NumberColumn(
                format="%.2f", disabled=True,
                help="Suma de las materias primas de la fila. Se recalcula al guardar."),
            "Balanza (TN)": st.column_config.NumberColumn(
                format="%.2f", disabled=True,
                help="Kg pesados en portería para este despacho (tickets vinculados)."),
            "Dif. vs balanza": st.column_config.NumberColumn(
                format="%.2f", disabled=True,
                help="Mezcla − balanza. Cerca de cero = los kg por materia prima cierran "
                     "con lo que salió de verdad.")}
    for c in _cols_mp:
        _cfg[c] = st.column_config.NumberColumn(c, format="%.2f", min_value=0.0, step=0.01,
                                                help="TN reales de %s. Vacío = vale lo "
                                                     "formulado." % c)
    _k = "dsw_ed_%s_%d" % (semana, int(st.session_state.get("dsw_ed_nonce") or 0))
    ed = st.data_editor(_vista, hide_index=True, use_container_width=True, key=_k,
                        column_config=_cfg,
                        disabled=["Aj.", "Despacho", "Fecha", "Cont", "Mezcla (TN)",
                                  "Balanza (TN)", "Dif. vs balanza"],
                        height=min(38 * (len(_vista) + 1) + 8, 460))

    g1, g2, g3 = st.columns([1.25, 1.5, 3.25])
    if g1.button("💾 Guardar los kg reales", type="primary", use_container_width=True,
                 key="dsw_guardar"):
        cambios = []
        for _i, idd in enumerate(_ids):
            for c in _cols_mp:
                try:
                    _new = ed.iloc[_i][c]
                    _old = base.iloc[_i][c]
                except Exception:
                    continue
                _vacio = _new is None or (isinstance(_new, float) and pd.isna(_new))
                if _vacio:
                    if str(base.iloc[_i]["Aj."]).startswith("✏️"):
                        cambios.append((idd, c, None))     # volver a lo formulado
                    continue
                if abs(float(_new) - float(_old or 0.0)) > 0.005:
                    cambios.append((idd, c, float(_new) * 1000.0))
        if not cambios:
            st.info("No cambiaste ningún valor.")
        elif conectar is None:
            st.error("Esta vista se abrió sin permisos de escritura.")
        else:
            try:
                n = _guardar_ajustes(conectar, USR, cambios)
                cat.clear()
                st.session_state["dsw_ed_nonce"] = int(
                    st.session_state.get("dsw_ed_nonce") or 0) + 1
                st.success("Guardado: %d ajuste(s). La tabla, el Excel y los PNG ya usan "
                           "estos kg." % n)
                st.rerun()
            except Exception as e:
                st.error("No se pudo guardar: %s" % e)

    _con_aj = [i for i, v in zip(_ids, base["Aj."]) if str(v).startswith("✏️")]
    if _con_aj and conectar is not None:
        _lbl = {int(r["id_despacho"]): str(r["Despacho"]) for _, r in base.iterrows()}
        _sel = g2.selectbox("Volver a lo formulado", ["—"] + [_lbl[i] for i in _con_aj],
                            key="dsw_reset_sel", label_visibility="collapsed")
        if _sel != "—" and g2.button("↩️ Borrar el ajuste de ese despacho",
                                     use_container_width=True, key="dsw_reset_go"):
            _idd = next(i for i in _con_aj if _lbl[i] == _sel)
            try:
                _guardar_ajustes(conectar, USR, [(_idd, m, None) for m in _cols_mp])
                cat.clear()
                st.session_state["dsw_ed_nonce"] = int(
                    st.session_state.get("dsw_ed_nonce") or 0) + 1
                st.success("Listo: %s vuelve a mostrar lo formulado." % _sel)
                st.rerun()
            except Exception as e:
                st.error("No se pudo borrar: %s" % e)
    g3.caption("El ajuste es por **materia prima**, no por tanque: los kg se reparten entre "
               "los tanques de esa materia prima en la misma proporción que tenía la "
               "formulación, así el detalle de líneas del Excel sigue cerrando.")


_CSS = """
<style>
/* Vista semanal: legibilidad de planta. Números grandes y tabulares, el color sólo
   cuando significa algo, y los ceros apagados para que no compitan con los valores. */
.dsw-hero{background:linear-gradient(135deg,#4f46e5 0%,#7c3aed 60%,#9333ea 100%);
  border-radius:18px;padding:18px 22px;color:#fff;margin:0 0 14px;position:relative;
  overflow:hidden;box-shadow:0 14px 34px -20px rgba(124,58,237,.75)}
.dsw-hero .g{position:absolute;right:-40px;top:-60px;width:190px;height:190px;
  background:radial-gradient(circle,rgba(255,255,255,.26),transparent 70%);pointer-events:none}
.dsw-hero h2{margin:0;color:#fff;font-size:1.22rem;font-weight:800;letter-spacing:-.01em}
.dsw-hero p{margin:5px 0 0;opacity:.93;font-size:.85rem;line-height:1.5;max-width:70ch}
.dsw-k{display:grid;grid-template-columns:repeat(auto-fit,minmax(150px,1fr));gap:12px;margin:0 0 14px}
.dsw-k .b{background:#fff;border:1px solid #e6e8f0;border-radius:15px;padding:13px 16px;
  box-shadow:0 1px 2px rgba(16,24,40,.05)}
.dsw-k .b.hi{background:linear-gradient(135deg,#eef2ff,#faf5ff);border-color:#e0e7ff}
.dsw-k .l{font-size:.66rem;text-transform:uppercase;letter-spacing:.07em;color:#64748b;font-weight:700}
.dsw-k .v{font-size:1.8rem;font-weight:800;color:#0f172a;line-height:1.1;margin-top:4px;
  font-variant-numeric:tabular-nums;font-family:'Plus Jakarta Sans','Inter',sans-serif}
.dsw-k .s{font-size:.75rem;color:#64748b;margin-top:2px}
.dsw-k .s.ok{color:#7c3aed;font-weight:700}
.dsw-t{width:100%;border-collapse:separate;border-spacing:0;background:#fff;
  border:1px solid #e6e8f0;border-radius:15px;overflow:hidden;font-variant-numeric:tabular-nums;
  box-shadow:0 1px 2px rgba(16,24,40,.05)}
.dsw-t th{background:#fafbff;border-bottom:1.5px solid #d7dbe8;padding:9px 11px;text-align:right;
  font-size:.66rem;font-weight:700;color:#64748b;text-transform:uppercase;letter-spacing:.07em;
  white-space:nowrap}
.dsw-t th.l{text-align:left}
.dsw-t td{padding:0 11px;height:46px;border-bottom:1px solid #eef0f6;text-align:right;
  font-size:.9rem;white-space:nowrap}
.dsw-t td.l{text-align:left}
.dsw-t tr:last-child td{border-bottom:0}
.dsw-t .d{font-weight:700}
.dsw-t .m{color:#64748b;font-size:.75rem}
.dsw-t .z{color:#a5aec0;font-weight:400}
.dsw-t .tt{font-weight:800;font-size:.98rem;font-family:'Plus Jakarta Sans','Inter',sans-serif}
.dsw-t .tot td{background:#f4f5fb;border-top:2px solid #d7dbe8;height:50px;font-weight:800}
.dsw-t .tag{display:inline-block;background:#f3f0ff;color:#7c3aed;border-radius:5px;
  padding:0 6px;font-size:.62rem;font-weight:800;margin-left:6px;vertical-align:1px}
@media (max-width:740px){ .dsw-t td,.dsw-t th{padding:0 6px;font-size:.8rem} }
</style>
"""


_COLS_TXT = ("Despacho", "Semana", "Fecha", "Rango", "Cliente", "Destino", "Estado")
_COLS_INT = ("Cont", "Despachos", "Contenedores")


def _tabla_html(t, mp_cols):
    """Tabla como HTML: ceros apagados, conteos sin decimales y el total con jerarquía.

    El TEXTO es el caso general y el número el particular: con la lista blanca al revés,
    columnas como Semana o Rango caían al formateador numérico, float("2026-S35")
    explotaba y la celda quedaba vacía."""
    cols = [c for c in t.columns if c not in ("Producto", "_aj")]
    h = "".join('<th class="l">%s</th>' % c if c in _COLS_TXT
                else "<th>%s</th>" % c for c in cols)
    _pri = next((c for c in cols if c in _COLS_TXT), cols[0] if cols else None)
    filas = []
    for i, r in t.iterrows():
        _tot = str(r.get(_pri) or "").startswith("TOTAL")
        tds = []
        for c in cols:
            v = r[c]
            if c in _COLS_INT:
                tds.append('<td>%s</td>'
                           % ("" if (v is None or v == "" or pd.isna(v)) else int(v)))
                continue
            _num = None
            if c not in _COLS_TXT:
                try:
                    _num = float(v)
                except Exception:
                    _num = None
            if _num is not None and not pd.isna(_num):
                _cl = "tt" if c == "TN total" else ("z" if abs(_num) < 0.005 else "")
                tds.append('<td class="%s">%s</td>' % (_cl, _fmt(_num)))
                continue
            _txt = "" if (v is None or (isinstance(v, float) and pd.isna(v))) else str(v)
            if c == _pri and _txt.startswith("TOTAL") is False and bool(r.get("_aj", False)):
                _txt += '<span class="tag" title="Este despacho tiene los kg reales '\
                        'cargados a mano">KG REALES</span>'
            tds.append('<td class="l%s">%s</td>' % (" d" if c == _pri else " m", _txt))
        filas.append('<tr class="%s">%s</tr>' % ("tot" if _tot else "", "".join(tds)))
    return '<table class="dsw-t"><thead><tr>%s</tr></thead><tbody>%s</tbody></table>' % (
        h, "".join(filas))


def render(USR, cat, conectar=None):
    st.markdown(_CSS, unsafe_allow_html=True)
    st.markdown(
        '<div class="dsw-hero"><div class="g"></div>'
        '<h2>📦 Despachos por semana</h2>'
        '<p>Cada semana con sus despachos y las materias primas en TN. Los que tengan '
        '<b>kg reales cargados</b> se muestran con esos kg; el resto, con los formulados. '
        'Abajo se descargan en Excel y PNG, y se ajustan a lo que realmente se cargó.</p>'
        '</div>', unsafe_allow_html=True)
    f1, f2, f3 = st.columns([1, 1.4, 1.6])
    sem = int(f1.number_input("Semanas hacia atrás", min_value=1, max_value=52, value=8, step=1, key="dsw_sem"))
    est = f2.multiselect("Estados", ["CONFIRMADO", "DESPACHADO", "BORRADOR", "ANULADO"], default=ESTADOS_DEF,
                         key="dsw_est")
    if not est:
        st.info("Elegí al menos un estado.")
        return
    try:
        raw = _lineas(cat, est, sem * 7 + 6)
    except Exception as e:
        st.error("No se pudieron leer los despachos: %s" % e)
        return
    if raw is None or raw.empty:
        st.info("No hay despachos con líneas en ese rango.")
        return
    df = _preparar(raw)
    # Los kg REALES cargados a mano pisan a los formulados ANTES de cualquier cálculo:
    # de acá para abajo la tabla, el resumen, el Excel y los PNG hablan de lo efectivo.
    try:
        _aj = _ajustes(cat)
        df = _aplicar_ajustes(df, _aj)
        df["tn"] = df["kg"] / 1000.0
    except Exception as _e:
        _aj = None
        df["ajustado"] = False
        st.warning("No se pudieron leer los ajustes de kg reales (%s): se muestra lo "
                   "formulado." % _e)
    try:
        _tk = _tickets(cat, sem * 7 + 6)
    except Exception:
        _tk = pd.DataFrame(columns=["id_despacho", "kg_tickets", "n_tickets", "n_contenedores"])
    _n_aj = int(df[df.get("ajustado", False) == True]["id_despacho"].nunique()) if "ajustado" in df else 0
    mps_all = _orden_mp(df)
    selp = f3.multiselect("Materias primas", mps_all, default=mps_all, key="dsw_mp")
    if selp:
        df = df[df["mp"].isin(selp)]
    if df.empty:
        st.info("Sin líneas con esas materias primas.")
        return
    mps = [m for m in mps_all if m in df["mp"].unique()]

    # ---- resumen + descargas globales
    res = _tabla_resumen(df, mps)
    st.markdown(
        '<div class="dsw-k">'
        '<div class="b"><div class="l">Semanas</div><div class="v">%d</div></div>'
        '<div class="b"><div class="l">Despachos</div><div class="v">%d</div></div>'
        '<div class="b"><div class="l">Contenedores</div><div class="v">%d</div></div>'
        '<div class="b hi"><div class="l">TN totales</div><div class="v">%s</div>'
        '<div class="s%s">%s</div></div>'
        '</div>' % (int(len(res) - 1), int(res["Despachos"].iloc[-1]),
                    int(res["Contenedores"].iloc[-1]), _fmt(res["TN total"].iloc[-1]),
                    " ok" if _n_aj else "",
                    ("%d con kg reales" % _n_aj) if _n_aj else "todo formulado"),
        unsafe_allow_html=True)
    with st.expander("📅 Resumen por semana", expanded=False):
        st.markdown(_tabla_html(res, mps + ["TN total"]), unsafe_allow_html=True)

    with st.spinner("Armando Excel y láminas…"):
        try:
            xlsx, zipb = _paquete(df, tuple(mps))
        except Exception as e:
            st.error("No se pudo armar el paquete: %s" % e)
            return
    with st.expander("⬇️ Descargar todo (Excel · PNG · ZIP)", expanded=False):
        d1, d2, d3 = st.columns(3)
        d1.download_button("Excel · todas las semanas", xlsx, file_name="despachos_semanal.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           key="dsw_xlsx", use_container_width=True, type="primary")
        d2.download_button("PNG · resumen de semanas", _png_resumen(df, mps),
                           file_name="resumen_semanas.png", mime="image/png",
                           key="dsw_png_res", use_container_width=True)
        d3.download_button("ZIP · Excel + PNG de cada semana", zipb,
                           file_name="despachos_semanal.zip", mime="application/zip",
                           key="dsw_zip", use_container_width=True)

    # ---- una semana
    st.markdown("<div style='height:6px'></div>", unsafe_allow_html=True)
    semanas = sorted(df["Semana"].unique().tolist(), reverse=True)
    c1, c2 = st.columns([1.5, 2.5])
    s = c1.selectbox("Semana", semanas, key="dsw_semana",
                     format_func=lambda x: "%s · %s" % (x, df[df["Semana"] == x]["Rango"].iloc[0]))
    df_s = df[df["Semana"] == s]
    t, mp_cols = _tabla_semana(df_s, mps)
    png = _png_semana(df_s, mps, s, df_s["Rango"].iloc[0])
    _nd = int(df_s["id_despacho"].nunique())
    _nc = int(df_s.groupby("id_despacho")["n_contenedores"].first().sum())
    c2.markdown(
        "<div style='padding-top:30px;color:#64748b;font-size:.86rem'>"
        "<b style='color:#0f172a;font-size:1.05rem'>%d despachos</b> · %d contenedores · "
        "<b style='color:#0f172a'>%s TN</b></div>" % (_nd, _nc, _fmt(float(df_s["tn"].sum()))),
        unsafe_allow_html=True)
    st.markdown(_tabla_html(t, mp_cols), unsafe_allow_html=True)
    st.download_button("⬇️ PNG de la semana %s" % s, png, file_name="despachos_%s.png" % s,
                       mime="image/png", key="dsw_png_sem")

    # ---- ajuste a los kg reales de esa semana ----
    if conectar is not None:
        with st.expander("✏️ Ajustar a los kg realmente cargados (semana %s)" % s,
                         expanded=False):
            try:
                _r = False
                if not st.session_state.get("dsw_clasico"):
                    _r = _grilla_reales(df_s, mps, _tk, _aj, cat, conectar, USR, s)
                if _r is not True:
                    _editor_reales(df_s, mps, _tk, cat, conectar, USR, s)
            except Exception as _e:
                import traceback as _tb
                st.error("No se pudo abrir el editor: %s" % _e)
                st.code(_tb.format_exc())

    with st.expander("🖼️ Ver la lámina de la semana (la misma que se descarga)", expanded=False):
        st.image(png, use_container_width=True)
